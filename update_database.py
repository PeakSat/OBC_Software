import openpyxl
import requests
import os

# Google Sheets details
spreadsheet_id = "12m4Sq4CMnLB9nn6INxKUdSNSIlTJes_uTbg19RE9-X0"

# Construct the export URL for .xlsx format
xlsx_export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"

# Output file
output_file = "parameter_database.xlsx"

try:
    # Send GET request
    response = requests.get(xlsx_export_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Write the content to a file
        with open(output_file, "wb") as file:
            file.write(response.content)
        print(f"Excel file downloaded successfully as '{output_file}'")
    else:
        print(f"Failed to download Excel file. HTTP Status Code: {response.status_code}")
except Exception as e:
    print(f"An error occurred: {e}")


# Subsystem acronyms and their corresponding number offsets
subsystem_config = {
    "OBDH": 0,
    "COMMS": 819,
    "PAY": 1638,
    "ADCS": 2457,
    "EPS": 3276,
}

# Type encoding dictionary for the last 4 bits
type_encoding = {
    "uint8_t": 0,
    "bool" : 0,
    "int8_t": 1,
    "uint16_t": 2,
    "int16_t": 3,
    "uint32_t": 4,
    "int32_t": 5,
    "uint64_t": 6,
    "int64_t": 7,
    "float": 8,
    "double": 9,
}

# Function to encode the ID
def encode_id(numeric_id, variable_type):
    if variable_type not in type_encoding:
        print(f"Error: Type '{variable_type}' not found in type encoding dictionary. Defaulting to 'uint64_t'.")
        variable_type = "uint64_t"
    type_code = type_encoding[variable_type]
    return (numeric_id << 4) | type_code

# Paths to the Excel file and output directories
excel_file = "parameter_database.xlsx"
inc_folder = "inc/Platform/Parameters"
output_hhp_file = os.path.join(inc_folder, "PeakSatParameters.hpp")

# Ensure the output directories exist
os.makedirs(inc_folder, exist_ok=True)

# Set to track unique IDs and store parameter types
processed_ids = set()
parameter_types = {}  # Dictionary to store {parameter_name: variable_type}

# Load the workbook
workbook = openpyxl.load_workbook(excel_file)

# Lists to collect output lines
hhp_lines = []

# Add the header lines for the .hpp file
hhp_lines.append("#pragma once")

# Process each subsystem separately
namespace_blocks = {acronym: [] for acronym in subsystem_config.keys()}
valid_rows = []

# Collect all valid rows across all sheets
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        id_cell = row[0]  # First column
        variable_cell = row[4]  # Fifth column

        # Skip rows with no variable name
        if not variable_cell.value or not isinstance(variable_cell.value, str):
            continue

        # Check if the first column contains a valid ID
        if id_cell.value and isinstance(id_cell.value, str):
            valid_rows.append(row)

# Process all valid rows
for idx, row in enumerate(valid_rows):
    id_cell = row[0]  # First column
    type_cell = row[2]  # Third column
    variable_cell = row[4]  # Fifth column
    enum_items_cell = row[6]  # Seventh column
    value_cell = row[7]  # Eighth column

    # Identify the subsystem acronym
    for acronym, offset in subsystem_config.items():
        if id_cell.value.startswith(f"{acronym}-"):
            numeric_id = id_cell.value[len(acronym) + 1:]  # Remove acronym and '-'

            try:
                numeric_id = int(numeric_id) + offset
            except ValueError:
                print(f"Skipping invalid numeric ID: {id_cell.value}")
                continue

            # Skip duplicates
            if numeric_id in processed_ids:
                continue
            processed_ids.add(numeric_id)

            # Get variable name and type
            variable_name = variable_cell.value.strip()
            variable_type = type_cell.value.strip() if type_cell.value else "int"

            # Handle "_enum" type
            if variable_type.endswith("_enum"):
                base_param = variable_type[:-5]  # Remove "_enum"
                if base_param in parameter_types:
                    variable_type = parameter_types[base_param]
                else:
                    print(f"Error: Parameter '{base_param}' not found for '{variable_name}_enum'. Defaulting to 'uint32_t'.")
                    variable_type = "uint32_t"

            # Store the type in the parameter_types dictionary
            parameter_types[variable_name] = variable_type

            # Encode the ID with the new encoding rule
            encoded_id = encode_id(numeric_id, variable_type)

            # Add to the corresponding namespace block
            block_lines = namespace_blocks[acronym]
            block_lines.append(f"    const ParameterId {variable_name}ID = {encoded_id};")

            # Enum definitions (if type is "enum")
            if variable_type in {"uint8_t", "uint16_t", "uint32_t", "uint64_t",
                                 "int8_t", "int16_t", "int32_t", "int64_t"} and enum_items_cell.value:
                enum_lines = (
                    f"    enum {variable_name}_enum : {variable_type} {{\n        {enum_items_cell.value.strip()}\n    }};"
                )
                block_lines.append(enum_lines)

            break

# Build the .hpp file
hhp_lines.append(f"namespace PeaksatParameters {{")
for acronym, block_lines in namespace_blocks.items():
    if block_lines:
        hhp_lines.extend(
            line for line in block_lines if not line.startswith("        ")
        )

hhp_lines.append("}")

# Add footer to the .hpp file

# Write the .hpp file
with open(output_hhp_file, "w") as hhp_file:
    hhp_file.write("\n".join(hhp_lines))

print(f"Processing complete.")
print(f"HHP file saved as '{output_hhp_file}'.")
